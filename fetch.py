#!/usr/bin/python
# -*- coding: utf-8 -*-

import pymysql.cursors
from openpyxl import load_workbook
from slugify import slugify
import re
from unicode_tr import unicode_tr
connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='bayhan',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
cursor = connection.cursor()

def titlecase(s):
    return re.sub(re.compile(r"[\w]+('[\w]+)?", flags=re.UNICODE),
                  lambda mo: mo.group(0)[0].upper() +
                             mo.group(0)[1:].lower(),
                  s)

try:
    wb = load_workbook('pk_2018_11_06.xlsx')
    ws = wb.active

    regions = {}
    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        index_str = str(index + 1)

        city = ws["A" + index_str].value
        district = ws["B" + index_str].value
        neighborhood = ws["C" + index_str].value
        part = ws["D" + index_str].value
        postal_code = ws["E" + index_str].value

        if city not in regions:
            regions[city] = {}

        if district not in regions[city]:
            regions[city][district] = {}

        if neighborhood not in regions[city][district]:
            regions[city][district][neighborhood] = {}

        regions[city][district][neighborhood][part] = postal_code

    for city in regions:
        #city_name = city.title().strip()
        city_name = unicode_tr(city.strip()).title()
        city_slug = slugify(city_name)



        with connection.cursor() as cursor:
            sql = "INSERT INTO `city` (`name`, `slug`) VALUES (%s, %s);"
            cursor.execute(sql, (city_name, city_slug))
            connection.commit()
            city_id = cursor.lastrowid

            for district in regions[city]:
                district_name = unicode_tr(district.strip()).title()
                district_slug = slugify(district_name)

                with connection.cursor() as cursor:
                    sql = "INSERT INTO `district` (`city_id`, `name`, `slug`) VALUES (%s, %s, %s);"
                    cursor.execute(sql, (city_id, district_name, district_slug))
                    connection.commit()
                    district_id = cursor.lastrowid

                    for neighborhood in regions[city][district]:
                        neighborhood_name = unicode_tr(neighborhood.strip()).title()
                        neighborhood_slug = slugify(neighborhood_name)

                        with connection.cursor() as cursor:
                            sql = "INSERT INTO `neighborhood` (`district_id`,`city_id`,`name`, `slug`) VALUES (%s,%s, %s, %s);"
                            cursor.execute(sql, (district_id,city_id,neighborhood_name, neighborhood_slug))
                            connection.commit()
                            neighborhood_id = cursor.lastrowid

                            for part in regions[city][district][neighborhood]:
                                part_name = unicode_tr(part.strip()).title()
                                part_slug = slugify(part_name)

                                with connection.cursor() as cursor:
                                    sql = "INSERT INTO `part` (`neighborhood_id`,`district_id`, `name`, `slug`, `postal_code`) VALUES (%s,%s, %s, %s, %s);"
                                    cursor.execute(sql, (neighborhood_id,district_id, part_name, part_slug, regions[city][district][neighborhood][part]))
                                    connection.commit()
                                    part_id = cursor.lastrowid
                                    postal_code = regions[city][district][neighborhood][part]

                                    print (city_name + ' / ' + district_name + ' / ' + neighborhood_name + ' / ' + part_name + ' - ' + str(postal_code))

finally:
    connection.close()
